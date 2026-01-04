import os
import re
import cv2
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
pytesseract.pytesseract.tesseract_cmd = r"C:\Tesseract\tesseract.exe"

DATA_DIR = r"C:\NLP-CV\dataset\CIN"   # optionally: ...\CIN\front and ...\CIN\back
OUT_XLSX = r"C:\NLP-CV\dataset\cin_front_back_2cols.xlsx"

VALID_EXT = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}

CFG_FRA = r"--oem 1 --psm 6"
CFG_ARA = r"--oem 1 --psm 6"
CFG_MRZ = r"--oem 1 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789<"

# =========================
# HELPERS
# =========================
def ocr(img, lang, cfg):
    t = pytesseract.image_to_string(img, lang=lang, config=cfg) or ""
    t = re.sub(r"\s{2,}", " ", t.replace("\n", " ")).strip()
    return t

# =========================
# EXCEL INIT (2 COLS)
# =========================
wb = Workbook()
ws = wb.active
ws.title = "dataset"

# For mDeBERTa classification:
# input  = text (OCR merged)
# output = label (front/back)
ws.append(["input", "output"])

ws["A1"].font = Font(bold=True)
ws["B1"].font = Font(bold=True)
ws.freeze_panes = "A2"

row_count = 0

for root, _, files in os.walk(DATA_DIR):
    parent = os.path.basename(root).lower()

    for fn in files:
        ext = os.path.splitext(fn)[1].lower()
        if ext not in VALID_EXT:
            continue

        path = os.path.join(root, fn)
        img = cv2.imread(path)
        if img is None:
            print("[SKIP] unreadable:", path)
            continue

        # ----- preprocess -----
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        enhanced = cv2.createCLAHE(2.0, (8, 8)).apply(gray)
        denoised = cv2.medianBlur(enhanced, 3)
        
        H, W = denoised.shape
        latin = cv2.adaptiveThreshold(denoised, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
        soft = denoised

        # ----- OCR texts -----
        text_fra = ocr(latin, "fra", CFG_FRA)
        text_ara = ocr(soft,  "ara", CFG_ARA)
        merged = (text_fra + " " + text_ara).strip()

        # ----- label front/back -----
        # If you have folders front/back, use them. Otherwise use MRZ heuristic.
        lower = latin[int(0.55 * H) : H, 0:W]
        mrz_text = ocr(lower, "eng", CFG_MRZ)
        has_mrz = ("<<" in mrz_text) or (re.search(r"<{2,}", mrz_text) is not None)

        if parent in {"front", "recto"}:
            label = "CIN_front"
        elif parent in {"back", "verso"}:
            label = "CIN_back"
        else:
            label = "CIN_back" if has_mrz else "CIN_front"

        # ----- write 2 columns -----
        ws.append([merged, label])

        row_count += 1
        print(f"[OK] {fn} | label={label}")

# ----- column widths -----
ws.column_dimensions[get_column_letter(1)].width = 80  # input
ws.column_dimensions[get_column_letter(2)].width = 12  # output

wb.save(OUT_XLSX)
print("\n✅ Excel saved:", OUT_XLSX)
print("Total samples:", row_count)
